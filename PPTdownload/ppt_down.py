# -*- coding: utf-8 -*-
# @Time    : 2020/7/27 11:30
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : ppt_down.py
# @Software: PyCharm

from selenium import webdriver
import time
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import os


def Downloadphoto(url,name,group):
    os.makedirs('./'+name+'/', exist_ok=True)
    from urllib.request import urlretrieve
    urlretrieve(url, './'+group+'/'+str(name)+'.jpg')




def PptDownUrl(driver,url,id):
    driver.get(url)

    try:
        Pdict = {'Pgroup': '', 'Pdownadr': ''}
        # pptg分组
        PgroupXpath = '//span[@class="c999 f12"]'

        # ppt下载地址
        PdownadrXpath = '//div[@id="url"]//a'

        WebDriverWait(driver, 10, 0.5).until(EC.visibility_of_all_elements_located((By.XPATH, PgroupXpath)))

        # pptg分组名称
        Pgroup = driver.find_element_by_xpath(PgroupXpath).text
        print(Pgroup.split('：')[1])
        # ppt下载地址
        Pdownadr = driver.find_element_by_xpath(PdownadrXpath).get_attribute('href')
        print(Pdownadr)

        Pdict['Pgroup'] = Pgroup.split('：')[1]
        Pdict['Pdownadr'] = Pdownadr
        print(Pdict)

    except Exception as e:
        print('获取失败！%s' % e)
        raise e

    writeExcel(id, Pdict)




def readExcel():
    rb_new = load_workbook('pptlist.xlsx')
    sheet = rb_new['Sheet1']
    data_list = []
    for i in range(sheet.max_row - 1):
        Pname = sheet.cell(row=i + 2, column=2).value
        PnameImg = sheet.cell(row=i + 2, column=3).value
        Pdown = sheet.cell(row=i + 2, column=4).value
        status = sheet.cell(row=i + 2, column=5).value  # 取值excel中的字典
        Pgroup = sheet.cell(row=i + 2, column=6).value
        Pdownadr = sheet.cell(row=i + 2, column=7).value
        dict = {'id': i + 1, 'Pname': Pname, 'PnameImg': PnameImg, 'Pdown': Pdown, 'status': status, 'Pgroup': Pgroup,
                'Pdownadr': Pdownadr}
        data_list.append(dict)
    return data_list



def writeExcel(id, dict):
    for i in range(1, len(dict) + 1):
        # 写入用例返回结果
        try:
            wb_new = load_workbook('pptlist.xlsx')
            sheet = wb_new['Sheet1']
            if 'Pgroup' in dict.keys():
                sheet.cell(id + 1, 6).value = dict['Pgroup']
                sheet.cell(id + 1, 7).value = dict['Pdownadr']
            if 'status' in dict.keys():
                sheet.cell(id + 1, 5).value = dict['status']
            wb_new.save('pptlist.xlsx')
        except Exception as e:
            print('数据写入失败！%s' % e)
            raise e
    print('id:{}执行写入excel成功！'.format(id))



def main():
    # 打开ppt网站页面
    driver = webdriver.Chrome()
    driver.maximize_window()

    list = readExcel()

    for k in range(len(list)):
        pdownUrl = list[k]['Pdown']
        # pnameImgUrl = list[k]['PnameImg']
        pdownId = list[k]['id']
        print(pdownUrl)

        PptDownUrl(driver,pdownUrl,pdownId)

    driver.quit()




if __name__ == '__main__':
    # 根据已有表格中下载页面获取下载地址，并写入表格
    # 已经获取就不需要执行
    main()






