# -*- coding: utf-8 -*-
# @Time    : 2020/7/27 9:18
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : ppt_spider.py
# @Software: PyCharm

from selenium import webdriver
import time
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook


def PPT(driver,url):

    driver.get(url)

    Plist = []

    for k in range(1, 21):
        try:
            Pdict = {'Pname': '', 'PnameImg': '', 'Pdown': '', 'status': 'X'}
            # 图片链接
            PoneXpath = '//div[@class="row"]//ul[@class="cards"]//div[' + str(k) + ']//img'

            # ppt下载页面地址
            PdownXpath = '//div[@class="row"]//ul[@class="cards"]//div[' + str(k) + ']//div[@class="download-now"]//a'

            WebDriverWait(driver, 10, 0.5).until(EC.visibility_of_all_elements_located((By.XPATH, PoneXpath)))

            # 图片名称
            Pname = driver.find_element_by_xpath(PoneXpath).get_attribute('alt')
            # print(Pname)
            # 图片地址
            PnameImg = driver.find_element_by_xpath(PoneXpath).get_attribute('src')
            # print(PnameImg)

            # ppt下载页面
            Pdown = driver.find_element_by_xpath(PdownXpath).get_attribute('href')
            # print(Pdown)

            Pdict['Pname'] = Pname
            Pdict['PnameImg'] = PnameImg
            Pdict['Pdown'] = Pdown
            print(Pdict)
            # 将一页信息导入列表
            Plist.append(Pdict)
        except:
            print('错误：', k)
            continue
    # time.sleep(1)
    # driver.quit()
    return Plist


def writeExcel(id, dict):
    id = id * 20 - 20
    for i in range(1, len(dict) + 1):
        # 写入用例返回结果
        try:
            wb_new = load_workbook('pptlist.xlsx')
            sheet = wb_new['Sheet1']
            sheet.cell(id + i + 1, 1).value = id + i
            sheet.cell(id + i + 1, 2).value = dict[i - 1]['Pname']
            sheet.cell(id + i + 1, 3).value = dict[i - 1]['PnameImg']
            sheet.cell(id + i + 1, 4).value = dict[i - 1]['Pdown']
            sheet.cell(id + i + 1, 5).value = dict[i - 1]['status']

            wb_new.save('pptlist.xlsx')
        except Exception as e:
            print('数据写入失败！%s' % e)
            raise e
    print('执行写入excel成功！')


def main():
    # 打开ppt网站页面
    driver = webdriver.Chrome()
    driver.maximize_window()

    for k in range(116, 152):
        if k == 1:
            url = 'https://www.52ppt.com/moban/index.html'
        else:
            url = 'https://www.52ppt.com/moban/index_' + str(k) + '.html'
        writeExcel(k, PPT(driver, url))
        print('第{}页完成写入！'.format(k))

    driver.quit()

if __name__ == '__main__':
    main()