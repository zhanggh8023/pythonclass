
# -*- coding: utf-8 -*-


from openpyxl import load_workbook
import os


class readExcel:

    def read_Excel(self,sheetname):
        rd = load_workbook('JGHG.xlsx')
        #print(rd.sheetnames)
        sheet = rd[sheetname]
        list = []
        list_2=[]
        dict={}

        for ii in range(sheet.max_column):
            name=sheet.cell(row=2,column=ii+1).value
            #print(name)
            list.append(name)
        #print(list)
            #print (ii)

        for i in range(1,sheet.max_row-1):
            for ii in range(sheet.max_column):
                name_value = sheet.cell(row=i+2, column=ii+1).value
                #print(name_value)
                #print(list[ii])
                #print(i)
                #print(ii)
                dict[list[ii]]=name_value
            dict.pop(sheetname)
            #print(dict)
            list_2.append(dict)

        return list_2


if __name__ == '__main__':
    print(readExcel().read_Excel('机构信息表YJPT_JGXXB'))
