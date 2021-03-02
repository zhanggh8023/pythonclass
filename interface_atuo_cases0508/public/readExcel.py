# -*- coding: utf-8 -*-
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : readExcel.py
# @Software: PyCharm


from public.config import config
from openpyxl import load_workbook
from conf import Allpath
from public.logger import Log

logger = Log('readExcel', Allpath.log_path)
s = config().read_config(Allpath.case_conf_path, 'SHEET', 'sheet')
url_1 = config().read_config(Allpath.http_conf_path, 'HTTP', 'url_f')
url_2 = config().read_config(Allpath.http_conf_path, 'HTTP', 'url_h')
url_3 = config().read_config(Allpath.http_conf_path, 'HTTP', 'url_a')
url_4 = config().read_config(Allpath.http_conf_path, 'HTTP', 'url_b')
mode = config().read_config(Allpath.case_conf_path, 'FLAG', 'mode')
logger.info("用例执行模式（0全部执行，1指定接口，2区间接口）：%s" % mode)
# logger.info("服务地址配置：{},{}".format(url_1, url_2, url_3))
case_list = config().read_config(Allpath.case_conf_path, 'FLAG', 'case_list')
case_list1 = config().read_config(Allpath.case_conf_path, 'FLAG', 'case_list1')
logger.info("配置用例列表：指定{},区间{}".format(case_list, case_list1))


class Case_Id:
    a, b = case_list1[0], case_list1[1]
    c = []
    if a > 0 or b > 0:
        if a <= b:
            for i in range(a, b + 1):
                c.append(i)
    if len(c) > 0:
        case_list1 = c
    else:
        case_list1 = case_list


class readExcel:
    def __init__ (self, file, sheet):
        self.file = file
        self.sheet = sheet

    def read_Excel (self):
        rd = load_workbook(self.file)
        sheet = rd[self.sheet]
        logger.info('获取表单成功：%s' % sheet)

        # initphone=self.get_init_phone()
        # print(initphone)#当作一个局部变量来看这个，一轮循环下来后在更新到excel
        # 0全部，1指点，2指定区间
        data_list = []
        if mode == 0:
            for i in range(sheet.max_row - 1):
                # print(i)
                method = sheet.cell(row=i + 2, column=2).value
                case_name = sheet.cell(row=i + 2, column=3).value
                env = eval(sheet.cell(row=i + 2, column=4).value)
                url = str(sheet.cell(row=i + 2, column=5).value)
                data = str(sheet.cell(row=i + 2, column=6).value)  # 取值excel中的字典
                sql = eval(sheet.cell(row=i + 2, column=7).value)
                assert_value = eval(sheet.cell(row=i + 2, column=8).value)
                code = str(sheet.cell(row=i + 2, column=9).value)
                if 'env' in env.keys() and env['env'] == 1:
                    url = url_2 + url
                elif 'env' in env.keys() and env['env'] == 2:
                    url = url_3 + url
                elif 'env' in env.keys() and env['env'] == 3:
                    url = url_4 + url
                else:
                    url = url_1 + url
                dict = {'id': i + 1, 'case_name': case_name, 'url': url, 'method': method, 'sql': sql, 'data': data,
                        'code': code, 'env': env, 'assert_value': assert_value}
                data_list.append(dict)  # self.update_phone(initphone)#把最后的值更新到excel中

        if mode == 1:
            for i in range(sheet.max_row - 1):
                id = sheet.cell(row=i + 2, column=1).value
                if id in case_list:
                    # print(i)
                    method = sheet.cell(row=i + 2, column=2).value
                    case_name = sheet.cell(row=i + 2, column=3).value
                    env = eval(sheet.cell(row=i + 2, column=4).value)
                    url = str(sheet.cell(row=i + 2, column=5).value)
                    data = str(sheet.cell(row=i + 2, column=6).value)  # 取值excel中的字典
                    sql = eval(sheet.cell(row=i + 2, column=7).value)
                    assert_value = eval(sheet.cell(row=i + 2, column=8).value)
                    code = str(sheet.cell(row=i + 2, column=9).value)
                    if 'env' in env.keys() and env['env'] == 1:
                        url = url_2 + url
                    elif 'env' in env.keys() and env['env'] == 2:
                        url = url_3 + url
                    elif 'env' in env.keys() and env['env'] == 3:
                        url = url_4 + url
                    else:
                        url = url_1 + url
                    dict = {'id': i + 1, 'case_name': case_name, 'url': url, 'method': method, 'sql': sql, 'data': data,
                            'code': code, 'env': env, 'assert_value': assert_value}
                    data_list.append(dict)
        if mode == 2:
            for i in range(sheet.max_row - 1):
                id = sheet.cell(row=i + 2, column=1).value
                if id in Case_Id.case_list1:
                    # print(i)
                    method = sheet.cell(row=i + 2, column=2).value
                    case_name = sheet.cell(row=i + 2, column=3).value
                    env = eval(sheet.cell(row=i + 2, column=4).value)
                    url = str(sheet.cell(row=i + 2, column=5).value)
                    data = str(sheet.cell(row=i + 2, column=6).value)  # 取值excel中的字典
                    sql = eval(sheet.cell(row=i + 2, column=7).value)
                    assert_value = eval(sheet.cell(row=i + 2, column=8).value)
                    code = str(sheet.cell(row=i + 2, column=9).value)
                    if 'env' in env.keys() and env['env'] == 1:
                        url = url_2 + url
                    elif 'env' in env.keys() and env['env'] == 2:
                        url = url_3 + url
                    elif 'env' in env.keys() and env['env'] == 3:
                        url = url_4 + url
                    else:
                        url = url_1 + url
                    dict = {'id': i + 1, 'case_name': case_name, 'url': url, 'method': method, 'sql': sql, 'data': data,
                            'code': code, 'env': env, 'assert_value': assert_value}
                    data_list.append(dict)

        # print(data_list)
        return data_list


# 根据id获取单行接口请求返回结果值
def read_test_result (id):
    file = Allpath.test_data_path
    rd = load_workbook(file)
    sheet = rd[s]
    value = str(sheet.cell(row=id + 1, column=13).value)
    logger.info('获取{}表单第{}条测试用例返回结果成功：{}'.format(s, id, value))
    return value


def readexcel ():
    file = Allpath.test_data_path
    sheet = 'Sheet2'
    rd = load_workbook(file)
    sheet = rd[sheet]
    logger.info('获取历史测试表单成功：%s' % sheet)
    dict = {'restult': '', 'sum': [], 'ok': [], 'fail': [], 'error': [], 'error_1': []}
    max = sheet.max_row + 1
    mix = sheet.max_row - 10
    if mix <= 1:
        mix = 2
    else:
        mix = max - 10
    for i in range(mix, max):
        dict['restult'] = sheet.cell(row=2, column=1).value
        dict['sum'].append(sheet.cell(row=i, column=2).value)
        dict['ok'].append(sheet.cell(row=i, column=3).value)
        dict['fail'].append(sheet.cell(row=i, column=4).value)
        dict['error'].append(sheet.cell(row=i, column=5).value)
        dict['error_1'].append(sheet.cell(row=i, column=5).value)
    return dict


if __name__ == '__main__':
    readexcel()
    readExcel(Allpath.test_data_path, 'Sheet1').read_Excel()
