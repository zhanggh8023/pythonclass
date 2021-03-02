# -*- coding: utf-8 -*-
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : writeExcel.py
# @Software: PyCharm


from openpyxl import load_workbook
from conf import Allpath
from public.logger import Log

logger = Log('writeExcel', Allpath.log_path)


class writeExcel:
    def __init__ (self, file, sheet):
        self.file = file
        self.sheet = sheet

    def write_Excel (self, i, dict):
        # 写入用例返回结果
        try:
            wb_new = load_workbook(self.file)
            sheet = wb_new[self.sheet]
            sheet.cell(i, 10).value = dict['code']
            sheet.cell(i, 11).value = dict['sql_result']
            sheet.cell(i, 12).value = dict['result']
            if 'data' in dict.keys():
                sheet.cell(i, 13).value = str(dict['data'])
            sheet.cell(i, 14).value = dict['time']
            wb_new.save(self.file)
            logger.info('执行写入excel成功！')
        except Exception as e:
            logger.info('数据写入失败！%s' % e)
            raise e

    def Excel_dellog (self):
        # 清除原始记录
        wb_new = load_workbook(self.file)
        sheet = wb_new[self.sheet]
        for i in range(sheet.max_row - 1):
            sheet.cell(row=i + 2, column=10).value = ""
            sheet.cell(row=i + 2, column=11).value = ""
            sheet.cell(row=i + 2, column=12).value = ""
            sheet.cell(row=i + 2, column=13).value = ""
            sheet.cell(row=i + 2, column=14).value = ""
        wb_new.save(self.file)
        logger.info('清除excel原始记录成功！')

    def write_time (self, i, dict):
        try:
            wb_new = load_workbook(self.file)
            sheet = wb_new[self.sheet]
            sheet.cell(i + 1, 15).value = dict['timemax']
            sheet.cell(i + 1, 16).value = dict['timemin']
            sheet.cell(i + 1, 17).value = dict['timesum']
            sheet.cell(i + 1, 18).value = dict['timemean']
            sheet.cell(i + 1, 19).value = dict['failsum']
            wb_new.save(self.file)
            logger.info('执行单接口性能数据写入excel成功！')
        except Exception as e:
            logger.info('性能数据写入失败！%s' % e)
            raise e


if __name__ == '__main__':
    writeExcel.write_time(1, {'timemax': 1.270276, 'timemin': 0.564675, 'timesum': 7.786537, 'timemean': '0.778654',
                              'failsum': '0.00'})
