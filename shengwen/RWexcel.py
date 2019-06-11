# -*- coding: utf-8 -*-
# @Time    : 2019/6/6 4:33
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : RWexcel.py
# @Software: PyCharm


from openpyxl import load_workbook

wb = load_workbook('voicedata.xlsx')


def readexcel(sheet):
    data = []

    sheet = wb[sheet]
    for i in range(sheet.max_row - 1):
        name = sheet.cell(row=i + 2, column=1).value
        phone = sheet.cell(row=i + 2, column=2).value

        dict = {'id': i + 1, 'name': name, 'phone': phone}
        data.append(dict)
    print('读取数据成功：', data)
    return data


def write_Excel(file, sheet, i, dcit, l):
    try:
        wb_new = load_workbook(file)
        sheet = wb_new[sheet]
        for ii in range(l):
            sheet.cell(i, ii + 3).value = dcit[ii]
        wb_new.save(file)
        print('执行写入excel成功！')
    except Exception as e:
        print('数据写入失败！%s' % e)
        raise e


def write_sqldata(file, sheet, i, dcit):
    try:
        wb_new = load_workbook(file)
        sheet = wb_new[sheet]
        for ii in range(i):
            sheet.cell(ii + 2, 1).value = dcit[ii]['name']
            sheet.cell(ii + 2, 2).value = dcit[ii]['phone']
            sheet.cell(1, ii + 3).value = dcit[ii]['name']
        wb_new.save(file)
        print('执行写入excel成功！')
    except Exception as e:
        print('数据写入失败！%s' % e)
        raise e


if __name__ == '__main__':
    readexcel('Sheet1')
    write_Excel('voicedata.xlsx', 'Sheet1', 2,
                ['1', '2', '3', '4', '1', '2', '3', '4', '1', '2', '3', '4', '1', '2', '3', '4', '1', '2', '3', '4',
                 '1', '2', '3'])
    write_sqldata('voicedata.xlsx', 'Sheet1', 2, )
