# -*- coding: utf-8 -*-
# @Time    : 2019/3/25 19:29
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : 跨表数据添加.py
# @Software: PyCharm


from openpyxl import load_workbook


# 默认可读写，若有需要可以指定write_only和read_only为True
wb = load_workbook('数据.xlsx')

# 获得所有sheet的名称
print(wb.sheetnames)
# 根据sheet名字获得sheet
a_sheet = wb['总装']
# 获得sheet名
print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = wb.defined_names