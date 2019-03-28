# -*- coding: utf-8 -*-
# @Time    : 2019/3/25 21:17
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : exls数据查询插入.py
# @Software: PyCharm

from openpyxl import load_workbook

wb= load_workbook('数据.xlsx')

data=[]
data1=[]

def red_GD():
    sheet = wb['工单拆分']
    for i in range(3,sheet.max_row):
        number_id=sheet.cell(row=i,column=7).value
        # number_id=sheet['G3'].value
        data1.append(number_id)


def red_ZZ():
    sheet = wb['总装']
    for i in range(sheet.max_row):
        for ii in range(sheet.max_column):
            dd=sheet.cell(row=i+3,column=ii+1).value
            # print(dd)
            dict={'i':i+3,'data':dd}
            # print(dict)
            data.append(dict)

red_ZZ()
red_GD()


for ii in range(len(data1)):
    for i in range(len(data)):
        sheet1=wb['总装']
        sheet = wb['日计划量']
        ff=data[i]['data']
        if data1[ii]==ff:
            tt=sheet1.cell(row=data[i]['i'],column=2).value
            print(ff,data1[ii])
            print(str(ii)+'-------------------对比正确班组：'+tt)
            sheet.cell(ii + 3, 6).value =tt
            break
wb.save('数据.xlsx')
print(data,data1)

