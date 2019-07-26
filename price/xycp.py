import mysql.connector
from zghmysql.db_config import readConfig
from zghmysql.readexcel import readExcel

#连接数据库
#1、数据库连接信息
config={'host':'47.110.127.36',
        'user':'root',
        'password':'citydo@123',
        'port':3306,
        'database':'xycp',
        }

#2、登录数据库
connector=mysql.connector.connect(**config)#  **后面加上配置文件名

#调用配置文件模块读取配置信息进行登录
# connector=mysql.connector.connect(**readConfig().read_config('db.conf','DBCONFIG','config'))

#建游标
cursor=connector.cursor(buffered=True)

#新建数据库
#cursor.execute("CREATE DATABASE zgh1practice")

#删除数据库
#cursor.execute("drop database zgh_practice")

#cursor.execute(zgh)#新建表单
#cursor.execute(sql_create_table)


#cursor.execute('drop table zgh_practice')#删除表单
list_1={}
def auto_insert(table):
    #执行SQL语句
    # cur= cursor.execute("SELECT DISTINCT TABLE_NAME FROM infomation_schema.COLUMNS WHERE COLUMN_NAME = loan." + table )
    # cur = cursor.execute("select ZGXM,SFZH from d_sgjj_grdkyqxx")

    # 传入sql语句，对应字段值
    cursor.execute("select ZGXM,SFZH from d_sgjj_grdkyqxx")
    #查看数据库数据
    data = cursor.description

    result = cursor.fetchall()

    print(result)

    data_1=[]
    test=[]
    for i in range(len(data)):
        data_1.append(data[i][0])
        print(data[i][0])
        test.append('%('+data[i][0]+')s')#生成value数列
        list_1[data[i][0]]=''
    print(data_1)
    # table_sql="INSERT INTO "+ table + "("+ ','.join(data_1) + ")" + " VALUES " + "("+ ','.join(test) + ")"


    # data_2=[i for i in reversed(data_1)]#反序排列
    # print(','.join(data_1))  # 转为元组
    # print(','.join(test))#转为元组
    # print(data)
    # print(len(data))
    return result

# auto_insert('d_sgjj_grdkyqxx')


#切记一定要执行
cursor.execute('commit')



from openpyxl import Workbook,load_workbook

def write_Excel( i, dict):
    file = 'xycp.xlsx'
    sheet = 'Sheet1'
    print(dict)
    # 写入用例返回结果
    try:
        wb_new = load_workbook(file)
        sheet = wb_new[sheet]
        dict1=[]
        for ii in range(len(dict)):
            if dict[ii] not in dict1:
                dict1.append(dict[ii])
        for i in range(len(dict1)):
            sheet.cell(i+2, 1).value = dict1[i][0]
            # sheet.cell(i,9).value = dcit['sql_result']
            sheet.cell(i+2, 2).value = dict1[i][1]

        wb_new.save(file)
        print('执行写入excel成功！')
    except Exception as e:
        print('数据写入失败！%s' % e)

write_Excel(0,auto_insert('d_sgjj_grdkyqxx'))



#关闭游标
cursor.close()

#关闭连接
connector.close()


#读取数据
def readexcel():
    file = 'xycp.xlsx'
    sheet = 'Sheet1'
    rd = load_workbook(file)
    sheet = rd[sheet]
    print('获取历史测试表单成功：%s' % sheet)

    data=[]
    for i in range(2,sheet.max_row + 1):
        dict = {"serveNo": "1563432328318457114", 'name': '', 'identity': ''}
        dict['name'] = sheet.cell(row=i, column=1).value
        dict['identity']=(sheet.cell(row=i, column=2).value)
        data.append(dict)
    print(data)
    return data

readexcel()

import requests

url='http://47.97.152.55:8507/serve/start'
headers={"Content-Type": "application/json"}
data=readexcel()

#写入excel
file = 'xycp.xlsx'
sheet = 'Sheet1'
wb_new = load_workbook(file)
sheet = wb_new[sheet]
for i in range(len(data)):
    print('现在开始进行post请求')
    # body格式
    request = requests.post(url, json=data[i], headers=headers).json()
    print(request)
    value=eval(request['responseData'])['productTemplateVOList'][0]['value']
    print(value)

    try:
        sheet.cell(i + 2, 3).value = value
        print('执行写入excel成功！')
    except Exception as e:
        print('数据写入失败！%s' % e)

wb_new.save(file)