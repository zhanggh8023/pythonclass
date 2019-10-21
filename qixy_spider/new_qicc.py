# -*- coding: utf-8 -*-
# @Time    : 2019/10/19 23:10
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : new_qicc.py
# @Software: PyCharm

import json
import codecs
import time
import re
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from collections import OrderedDict
from openpyxl import load_workbook


def re_excel (file, sheetname):
    rd = load_workbook(file)
    sheet = rd[sheetname]
    data_list = []

    for i in range(sheet.max_row - 1):
        data = {}
        name = sheet.cell(row=i + 2, column=1).value
        namedate = sheet.cell(row=i + 2, column=2).value
        data["id"] = i
        data[i] = namedate
        data['name']=name

        data_list.append(data)
    print(data_list, len(data_list))
    return data_list


data = re_excel('enterprise_data.xlsx', 'Sheet1')


def wr_excel (dict):
    wb_new = load_workbook('enterprise_data.xlsx')
    sheet = wb_new['Sheet1']
    for i in range(len(data)):
        if (dict['统一社会信用代码'] == data[i][i])| (dict['企业名称']==data[i]['name']) | (dict['注册号'] == data[i][i]):
            sheet.cell(i + 2, 3).value = dict['企业名称']
            sheet.cell(i + 2, 4).value = dict['统一社会信用代码']
            sheet.cell(i + 2, 5).value = dict['注册号']
            sheet.cell(i + 2, 6).value = dict['组织机构代码']
            sheet.cell(i + 2, 7).value = dict['经营状态']
            sheet.cell(i + 2, 8).value = dict['公司类型']
            sheet.cell(i + 2, 9).value = dict['成立日期']
            sheet.cell(i + 2, 10).value = dict['法定代表人']
            sheet.cell(i + 2, 11).value = dict['注册资本']
            sheet.cell(i + 2, 12).value = dict['实缴资本']
            sheet.cell(i + 2, 13).value = dict['营业期限']
            sheet.cell(i + 2, 14).value = dict['登记机关']
            sheet.cell(i + 2, 15).value = dict['发照日期']
            sheet.cell(i + 2, 16).value = dict['公司规模']
            sheet.cell(i + 2, 17).value = dict['所属行业']
            sheet.cell(i + 2, 18).value = dict['英文名']
            sheet.cell(i + 2, 19).value = dict['曾用名']
            sheet.cell(i + 2, 20).value = dict['企业地址']
            sheet.cell(i + 2, 21).value = dict['经营范围']

            print('写入成功！')
            wb_new.save('enterprise_data.xlsx')
            pass



class Tianyancha:
    # 常量定义
    url = 'https://www.tianyancha.com/login'

    def __init__ (self, username, password, headless=False):
        self.username = username
        self.password = password
        self.headless = headless
        self.driver = self.login(text_login='请输入11位手机号码', text_password='请输入登录密码')

    # 登录天眼查
    def login (self, text_login, text_password):
        time_start = time.time()

        # 操作行为提示
        print('在自动输入完用户名和密码前，请勿操作鼠标键盘！请保持优雅勿频繁（间隔小于1分钟）登录以减轻服务器负载。')

        # 设置是否为隐藏加载并打开浏览器
        if self.headless:
            option = webdriver.ChromeOptions()
            option.add_argument('headless')
            driver = webdriver.Chrome(chrome_options=option)
        else:
            driver = webdriver.Chrome()

        # 强制声明浏览器长宽为1024*768以适配所有屏幕
        driver.set_window_position(0, 0)
        driver.set_window_size(1024, 768)
        driver.get(self.url)

        # 模拟登陆：Selenium Locating Elements by Xpath
        time.sleep(1)
        # 等待登录按钮元素出现
        WebDriverWait(driver, 10, 0.5).until(
            EC.visibility_of_all_elements_located((By.XPATH,"//img[@id='tyc_banner_close']")))
        # 关闭底栏
        driver.find_element_by_xpath("//img[@id='tyc_banner_close']").click()
        driver.find_element_by_xpath("//div[@tyc-event-ch='Login.PasswordLogin']").click()
        # 天眼查官方会频繁变化登录框的占位符,所以设置两个新参数来定义占位符
        driver.find_elements_by_xpath("//input[@placeholder='{}']".format(text_login))[-2].send_keys(self.username)
        driver.find_elements_by_xpath("//input[@placeholder='{}']".format(text_password))[-1].send_keys(self.password)

        # 手工登录，完成滑块验证码
        print('请现在开始操作键盘鼠标，在15s内点击登录并手工完成滑块验证码。批量爬取只需一次登录。')
        time.sleep(10)
        print('还剩5秒。')
        time.sleep(5)

        time_end = time.time()
        print('您的本次登录共用时{}秒。'.format(int(time_end - time_start)))
        return driver

    # 定义天眼查爬虫
    def tianyancha_scraper (self, keyword,change_page_interval=2,
                            export='xlsx', quit_driver=True):
        """
        天眼查爬虫主程序。
        :param keyword: 公司名称，支持模糊或部分检索。比如"北京鸿智慧通实业有限公司"。
        :param table: 需要爬取的表格信息，默认为全部爬取。和官方的元素名称一致。常见的可以是'baseInfo', 'staff', 'invest'等，具体请参考表格名称中英文对照表。
        :param use_default_exception: 是否使用默认的排除列表，以忽略低价值表格为代价来加快爬取速度。
        :param change_page_interval: 爬取多页的时间间隔，默认2秒。避免频率过快IP地址被官方封禁。
        :param export: 输出保存格式，默认为Excel的`xlsx`格式，也支持`json`。
        :return:
        """

        # 公司搜索：顺带的名称检查功能，利用天眼查的模糊搜索能力
        # TODO：将借用模糊搜索的思路写进宣传文章。
        def search_company (driver, url1):
            print(url1)
            driver.get(url1)
            time.sleep(1)
            content = driver.page_source.encode('utf-8')
            # TODO：是否要将登录状态监测统一到login函数中
            try:
                # TODO：'中信证券股份有限公司'无法正确检索
                try:
                    url1 = driver.find_element_by_xpath(
                        '//*[@id="web-content"]/div/div[1]/div[3]/div[2]/div[1]/div/div[3]/div[1]/a').get_attribute(
                        'href')
                except:
                    url1 = driver.find_element_by_xpath(
                        '//*[@id="web-content"]/div/div[1]/div[2]/div[2]/div/div/div[3]/div[1]/a').get_attribute(
                        'href')
                print('登陆成功。')
            except:
                print('登陆过于频繁，请1分钟后再次尝试。')




            # TODO: 如果搜索有误，手工定义URL2地址。有无改善方案？
            driver.get(url1)
            return driver


        def tryonclick (table):  # table实质上是selenium WebElement
            # 测试是否有翻页
            ## 把条件判断写进tryonclick中
            try:
                # 找到有翻页标记
                table.find_element_by_tag_name('ul')
                onclickflag = 1
            except Exception:
                print("没有翻页")  ## 声明表格名称: name[x] +
                onclickflag = 0
            return onclickflag

        def scrapy (driver):
            # # 强制确认table类型为list：当只爬取一个元素的时候很可能用户会只传入表明str
            # if isinstance(table, str):
            #     list_table = []
            #     list_table.append(table)
            #     table = list_table

            # # 定义排除列表
            # # TODO:允许用户自主选择保留项目;帮助检查没有重复项
            # if use_default_exception:
            #     list_exception = ['recruit', 'tmInfo', 'holdingCompany', 'invest', 'bonus', 'firmProduct', 'jingpin', \
            #                       'bid', 'taxcredit', 'certificate', 'patent', 'copyright', 'product', 'importAndExport', \
            #                       'copyrightWorks', 'wechat', 'icp', 'announcementcourt', 'lawsuit', 'court', \
            #                       'branch', 'touzi', 'judicialSale', 'bond', 'teamMember', 'check']
            #     # 两个List取差异部分，只排除不在爬取范围内的名单。参考：https://stackoverflow.com/questions/1319338/combining-two-lists-and-removing-duplicates-without-removing-duplicates-in-orig/1319353#1319353
            #     list_exception = list(set(list_exception) - set(table))
            # else:
            #     list_exception = []

            # Waiting time for volatilityNum to load
            # time.sleep(2)
            # tables = driver.find_elements_by_xpath("//div[contains(@id,'_container_')]")

            dict = {}
            try:
                # 等待登录按钮元素出现
                WebDriverWait(driver, 5, 0.5).until(
                    EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="company_web_top"]/div[2]/div[3]/div[1]/h1')))
                name = driver.find_element_by_xpath('//*[@id="company_web_top"]//div//h1').text
                dict['企业名称'] = name
                shxydm = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[3]/td[2]').text
                dict['统一社会信用代码'] = shxydm
                zch = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[3]/td[4]').text
                dict['注册号'] = zch
                zzjgdm = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[4]/td[4]').text
                dict['组织机构代码'] = zzjgdm
                jyzt = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[2]/td[4]').text
                dict['经营状态'] = jyzt
                gslx = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[5]/td[2]').text
                dict['公司类型'] = gslx
                clrq = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[6]/td[2]').text
                dict['成立日期'] = clrq
                fddbr = driver.find_element_by_xpath(
                    '//*[@id="_container_baseInfo"]/table[1]/tbody/tr[1]/td[1]/div/div[1]/div[2]/div[1]/a').text
                dict['法定代表人'] = fddbr
                zczb = driver.find_element_by_xpath(
                    '//*[@id="_container_baseInfo"]/table[2]/tbody/tr[1]/td[2]/div').get_attribute('title')
                print(zczb)
                dict['注册资本'] = zczb
                sjzb = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[1]/td[4]').text
                dict['实缴资本'] = sjzb
                yyqx = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[7]/td[2]/span').text
                dict['营业期限'] = yyqx
                djjg = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[6]/td[4]').text
                dict['登记机关'] = djjg
                fzrq = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[6]/td[2]').text
                dict['发照日期'] = fzrq
                gsgm = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[8]/td[2]').text
                dict['公司规模'] = gsgm
                sshy = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[5]/td[4]').text
                dict['所属行业'] = sshy
                ywm = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[9]/td[4]').text
                dict['英文名'] = ywm
                cym = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[9]/td[2]').text
                dict['曾用名'] = cym
                qydz = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[10]/td[2]').text
                dict['企业地址'] = qydz
                jyfw = driver.find_element_by_xpath('//*[@id="_container_baseInfo"]/table[2]/tbody/tr[11]/td[2]/span').text
                dict['经营范围'] = jyfw
                print(dict)
                wr_excel(dict)
            except:
                pass

        time_start = time.time()

        url_search = 'http://www.tianyancha.com/search?key=%s&checkFrom=searchBox' % keyword
        self.driver = search_company(self.driver, url_search)
        table_dict = scrapy(self.driver)
        # if export == 'xlsx':
        #     gen_excel(table_dict, keyword)
        # elif export == 'json':
        #     gen_json(table_dict, keyword)
        # else:
        #     print("请选择正确的输出格式，支持'xlsx'和'json'。")

        time_end = time.time()
        print('您的本次爬取共用时{}秒。'.format(int(time_end - time_start)))
        return table_dict

    # 定义批量爬取爬虫
    def tianyancha_scraper_batch (self, input_template='enterprise_data.xlsx', change_page_interval=2, export='xlsx'):
        df_input = pd.read_excel(input_template, encoding='gb18030').dropna(axis=1, how='all')
        list_dicts = []

        # 逐个处理输入信息
        for i in range(len(df_input)):
            print(i)
            keyword = df_input['提供代码匹配'].iloc[i]
            keyword1 = df_input['企业名称'].iloc[i]
            print(keyword,keyword1)
            if keyword1 == '空':

                # 批量调取天眼查爬虫
                table_dict = self.tianyancha_scraper(keyword=keyword,change_page_interval=change_page_interval, export=export,quit_driver=False)
                list_dicts.append(table_dict)
                print(table_dict)
            continue

        self.driver.quit()
        # 全部运行完后退出浏览器
        return tuple(list_dicts)


# 批量
tuple_dicts = Tianyancha(username='17681829051', password='zhang8023').tianyancha_scraper_batch(
    input_template='enterprise_data.xlsx', export='xlsx')
print(tuple_dicts[0])
