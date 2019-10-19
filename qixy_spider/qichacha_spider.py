# -*- coding: utf-8 -*-
# @Time    : 2019/10/18 23:52
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : qichacha_spider.py
# @Software: PyCharm

from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import os

rd = load_workbook('enterprise_data.xlsx')
sheet = rd['Sheet1']
data=[]

for i in range(sheet.max_row - 1):
    namedate = sheet.cell(row=i + 2, column=2).value
    data.append(namedate)
print(data,len(data))

driver=webdriver.Chrome()
driver.maximize_window()
driver.get('https://www.qichacha.com/')

driver.find_element_by_xpath('/html/body/header/div/ul/li[12]/a/span').click()

WebDriverWait(driver, 10, 0.5).until(EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="normalLogin"]')))

driver.find_element_by_xpath('//*[@id="normalLogin"]').click()
driver.find_element_by_xpath('//*[@id="nameNormal"]').send_keys('17681829051')
driver.find_element_by_xpath('//*[@id="pwdNormal"]').send_keys('123456')

# 起点
start = driver.find_element_by_id('nc_2_n1z')
# 终点
end = driver.find_element_by_xpath('//*[@id="user_login_normal"]/div[2]/div')

actions = ActionChains(driver)
actions.drag_and_drop(start, end)
time.sleep(1)
# 执行
actions.perform()
time.sleep(5)

pp='//*[@id="searchkey"]'
oo='//*[@id="V3_Search_bt"]'
uu='//*[@id="headerKey"]'
yy='/html/body/header/div/form/div/div/span/button'
hh='//*[@id="search-result"]/tr/td[3]/a'
dd='//*[@id="Cominfo"]/table'


dict_list={}


for ii in range(2):
    dict = {}
    print(data[ii])
    if ii <1:
        # 等待登录按钮元素出现
        WebDriverWait(driver, 10, 0.5).until(EC.visibility_of_all_elements_located((By.XPATH, pp)))
        # 输入输入内容
        driver.find_element_by_xpath(pp).send_keys(data[ii])
        driver.find_element_by_xpath(oo).click()
    else:
        driver.get('https://www.qichacha.com/search?key=370681018013864')
        #等待登录按钮元素出现
        WebDriverWait(driver,10,0.5).until(EC.visibility_of_all_elements_located((By.XPATH,uu)))
        driver.find_element_by_xpath(uu).clear()
        #输入输入内容
        driver.find_element_by_xpath(uu).send_keys(data[ii])
        driver.find_element_by_xpath(yy).click()

    #等待登录按钮元素出现
    # WebDriverWait(driver,10,0.5).until(EC.visibility_of_all_elements_located((By.XPATH,hh)))

    driver.find_element_by_xpath(hh).click()#点击进入
    time.sleep(1)

    windows = driver.window_handles
    # 切换好指定的窗口
    driver.switch_to.window(windows[-1])  # 最新打开的窗口
    #等待登录按钮元素出现
    WebDriverWait(driver,10,0.5).until(EC.visibility_of_all_elements_located((By.XPATH,'//*[@id="company-top"]//div//h1')))
    name = driver.find_element_by_xpath('//*[@id="company-top"]//div//h1').text
    dict['企业名称']=name
    shxydm=driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[4]/td[2]').text
    dict['统一社会信用代码']=shxydm
    zch=driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[5]/td[2]').text
    dict['注册号'] = zch
    zzjgdm = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[5]/td[4]').text
    dict['组织机构代码'] = zzjgdm
    jyzt = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[3]/td[2]').text
    dict['经营状态'] = jyzt
    gslx = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[6]/td[2]').text
    dict['公司类型'] = gslx
    clrq = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[9]/td[2]').text
    dict['成立日期'] = clrq
    fddbr = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[1]/td[2]/div/div/div[2]/a[1]/h2').text
    dict['法定代表人'] = fddbr
    zczb = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[1]/td[4]').text
    dict['注册资本'] = zczb
    yyqx = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[10]/td[4]').text
    dict['营业期限'] = yyqx
    djjg = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[7]/td[4]').text
    dict['登记机关'] = djjg
    fzrq = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[7]/td[2]').text
    dict['发照日期'] = fzrq
    gsgm = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[10]/td[2]').text
    dict['公司规模'] = gsgm
    sshy = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[6]/td[4]').text
    dict['所属行业'] = sshy
    ywm = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[8]/td[4]').text
    dict['英文名'] = ywm
    cym = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[9]/td[2]').text
    dict['曾用名'] = cym
    qydz = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[11]/td[2]').text
    dict['企业地址'] = qydz
    jyfw = driver.find_element_by_xpath('//*[@id="Cominfo"]/table/tbody/tr[12]/td[2]').text
    dict['经营范围'] = jyfw

    dict_list[ii]=dict
    driver.switch_to.window(windows[0])


    print(dict_list)


def write_Excel(i,dict):
     # 写入用例返回结果
    try:
        wb_new = load_workbook('enterprise_data.xlsx')
        sheet = wb_new['Sheet1']
        sheet.cell(i, 3).value = dict['公司名称']
        sheet.cell(i, 4).value = dict['统一社会信用代码']
        sheet.cell(i, 5).value = dict['注册号']
        sheet.cell(i, 6).value = dict['组织机构代码']
        sheet.cell(i, 7).value = dict['经营状态']
        sheet.cell(i, 8).value = dict['公司类型']
        sheet.cell(i, 9).value = dict['成立日期']
        sheet.cell(i, 10).value = dict['法定代表人']
        sheet.cell(i, 11).value = dict['注册资本']
        sheet.cell(i, 12).value = dict['营业期限']
        sheet.cell(i, 13).value = dict['登记机关']
        sheet.cell(i, 14).value = dict['发照日期']
        sheet.cell(i, 15).value = dict['公司规模']
        sheet.cell(i, 16).value = dict['所属行业']
        sheet.cell(i, 17).value = dict['英文名']
        sheet.cell(i, 18).value = dict['曾用名']
        sheet.cell(i, 19).value = dict['企业地址']
        sheet.cell(i, 20).value = dict['经营范围']

        wb_new.save('enterprise_data.xlsx')
        print('执行写入excel成功！')
    except Exception as e:
        print('数据写入失败！%s'%e)
        raise e